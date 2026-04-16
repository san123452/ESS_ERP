import io
import logging
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


logger = logging.getLogger(__name__)

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
XLSM_MIME = "application/vnd.ms-excel.sheet.macroEnabled.12"

MODULE_DIR = Path(__file__).resolve().parent
DEFAULT_TEMPLATE_PATH = MODULE_DIR / "templates" / "ESS_Report_Template.xlsm"
TEMPLATE_ENV_VAR = "ESS_REPORT_TEMPLATE_PATH"

SHEET_SUMMARY = "\uacbd\uc601_\uc694\uc57d"
SHEET_MONTHLY = "\uc6d4\ubcc4_\uc190\uc775"
SHEET_QUARTERLY = "\ubd84\uae30\ubcc4_\uc2e4\uc801"
SHEET_FORECAST = "AI_\uc7a5\uae30_\uc608\uce21"
SHEET_QUALITY = "\ud488\uc9c8_\ubd84\uc11d"
SHEET_CANCEL = "\ucde8\uc18c_\ubd84\uc11d"


class MissingMacroTemplateError(FileNotFoundError):
    pass


@dataclass
class ExcelReportArtifact:
    buffer: io.BytesIO
    filename: str
    media_type: str
    macro_enabled: bool
    template_path: Optional[Path] = None


P = {
    "navy": "1C365E",
    "navy_mid": "2B4E88",
    "green": "277C60",
    "green_lt": "D7F0E6",
    "red": "C0392B",
    "red_lt": "FAE0DC",
    "amber": "B47800",
    "amber_lt": "FFF4D2",
    "lt_gray": "F5F7FA",
    "mid_gray": "E5E9EF",
    "white": "FFFFFF",
    "text_main": "1A2742",
    "border": "C8D2E0",
}


def _thin_border() -> Border:
    side = Side(style="thin", color=P["border"])
    return Border(left=side, right=side, top=side, bottom=side)


def _medium_border() -> Border:
    side = Side(style="medium", color=P["navy"])
    return Border(left=side, right=side, top=side, bottom=side)


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _font(size: int = 10, bold: bool = False, color: Optional[str] = None, italic: bool = False) -> Font:
    return Font(name="Arial", size=size, bold=bold, italic=italic, color=color or P["text_main"])


CC = Alignment(horizontal="center", vertical="center")
CR = Alignment(horizontal="right", vertical="center")
CL = Alignment(horizontal="left", vertical="center")

HDR_FILL = _fill(P["navy"])
SUBHDR_FILL = _fill(P["navy_mid"])
ALT_FILL = _fill(P["lt_gray"])
TOT_FILL = _fill(P["mid_gray"])
NEG_FILL = _fill(P["red_lt"])
POS_FILL = _fill(P["green_lt"])
WARN_FILL = _fill(P["amber_lt"])
BTN_FILL = _fill(P["navy"])

HDR_FONT = _font(10, bold=True, color=P["white"])
BODY_FONT = _font(10)
BOLD_FONT = _font(10, bold=True)
TITLE_FONT = _font(14, bold=True, color=P["navy"])
FC_FONT = _font(10, italic=True, color=P["amber"])
NEG_FONT = _font(10, bold=True, color=P["red"])
BTN_FONT = _font(12, bold=True, color=P["white"])

NUM = "#,##0"
PCT = '0.00"%"'
NUM_N = '#,##0;[Red]-#,##0;"-"'


def build_excel_report(analysis_data: dict, require_macro_template: bool = False) -> ExcelReportArtifact:
    raw = analysis_data.get("data", analysis_data)
    template_path = _resolve_template_path()

    if template_path:
        workbook = load_workbook(template_path, keep_vba=True)
        artifact = _build_artifact(workbook, raw, macro_enabled=True, template_path=template_path)
        logger.info("Built macro-enabled report from template: %s", template_path)
        return artifact

    if require_macro_template:
        raise MissingMacroTemplateError(
            "Macro template not found. Place ESS_Report_Template.xlsm at "
            f"{DEFAULT_TEMPLATE_PATH} or set {TEMPLATE_ENV_VAR}."
        )

    logger.warning(
        "No xlsm template found. Returning xlsx without embedded macros. "
        "Set %s or place template at %s",
        TEMPLATE_ENV_VAR,
        DEFAULT_TEMPLATE_PATH,
    )
    workbook = Workbook()
    return _build_artifact(workbook, raw, macro_enabled=False, template_path=None)


def build_macro_enabled_report(analysis_data: dict) -> ExcelReportArtifact:
    return build_excel_report(analysis_data, require_macro_template=True)


def generate_excel_report(analysis_data: dict) -> io.BytesIO:
    return build_excel_report(analysis_data).buffer


def _build_artifact(
    workbook: Workbook,
    raw: dict,
    macro_enabled: bool,
    template_path: Optional[Path],
) -> ExcelReportArtifact:
    _reset_report_sheets(workbook)
    _build_summary(workbook, raw)
    _build_monthly(workbook, raw)
    _build_quarterly(workbook, raw)
    _build_forecast(workbook, raw)
    _build_quality(workbook, raw)
    _build_cancel(workbook, raw)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return ExcelReportArtifact(
        buffer=output,
        filename="ESS_ERP_Report.xlsm" if macro_enabled else "ESS_ERP_Report.xlsx",
        media_type=XLSM_MIME if macro_enabled else XLSX_MIME,
        macro_enabled=macro_enabled,
        template_path=template_path,
    )


def _resolve_template_path() -> Optional[Path]:
    env_path = os.getenv(TEMPLATE_ENV_VAR)
    if env_path:
        candidate = Path(env_path).expanduser()
        if candidate.exists():
            return candidate

    if DEFAULT_TEMPLATE_PATH.exists():
        return DEFAULT_TEMPLATE_PATH

    return None


def _reset_report_sheets(workbook: Workbook) -> None:
    for name in list(workbook.sheetnames):
        if name in {
            SHEET_SUMMARY,
            SHEET_MONTHLY,
            SHEET_QUARTERLY,
            SHEET_FORECAST,
            SHEET_QUALITY,
            SHEET_CANCEL,
        }:
            del workbook[name]

    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) == 1:
        del workbook["Sheet"]


def _vlen(value) -> int:
    return sum(2 if ord(ch) > 127 else 1 for ch in str(value))


def _auto_width(ws, max_width: int = 55) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max((_vlen(cell.value) for cell in col if cell.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max(width * 1.15 + 4, 13), max_width)


def _hdr_row(ws, row: int, ncols: int) -> None:
    border = _thin_border()
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CC
        cell.border = border


def _apply_alt(ws, row_index: int, ncols: int) -> None:
    if row_index % 2 == 0:
        for col in range(1, ncols + 1):
            ws.cell(row=row_index, column=col).fill = ALT_FILL


def _add_table(ws, name: str, ref: str, style: str = "TableStyleMedium9") -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name=style, showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)


def _cf_margin(ws, col: str, start: int, end: int) -> None:
    ws.conditional_formatting.add(
        f"{col}{start}:{col}{end}",
        CellIsRule(operator="lessThan", formula=["0"], font=NEG_FONT, fill=NEG_FILL),
    )
    ws.conditional_formatting.add(
        f"{col}{start}:{col}{end}",
        CellIsRule(operator="greaterThan", formula=["10"], font=_font(10, bold=True, color=P["green"]), fill=POS_FILL),
    )


def _cf_profit(ws, col: str, start: int, end: int) -> None:
    ws.conditional_formatting.add(
        f"{col}{start}:{col}{end}",
        CellIsRule(operator="lessThan", formula=["0"], fill=NEG_FILL),
    )


def _cf_databar(ws, col: str, start: int, end: int, color: str) -> None:
    ws.conditional_formatting.add(
        f"{col}{start}:{col}{end}",
        DataBarRule(start_type="min", start_value=0, end_type="max", end_value=None, color=color),
    )


def _safe_rows(values: Optional[Iterable]) -> list:
    return list(values or [])


def _build_summary(workbook: Workbook, raw: dict) -> None:
    ws = workbook.create_sheet(SHEET_SUMMARY)
    ws.sheet_properties.tabColor = P["navy"]

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    ws.merge_cells("A1:D2")
    cell = ws["A1"]
    cell.value = "Run Reports  (Enable Macro -> Click Button)"
    cell.font = BTN_FONT
    cell.fill = BTN_FILL
    cell.alignment = CC
    cell.border = _medium_border()

    ws.row_dimensions[3].height = 6
    ws.row_dimensions[4].height = 26
    ws.merge_cells("A4:D4")
    ws["A4"] = "ESS ERP \uacbd\uc601 \uc694\uc57d"
    ws["A4"].font = TITLE_FONT
    ws["A4"].alignment = CL

    headers = ["\uc9c0\ud45c", "\uac12", "\ub2e8\uc704", "\ube44\uace0"]
    for index, title in enumerate(headers, 1):
        ws.cell(row=5, column=index, value=title)
    _hdr_row(ws, 5, len(headers))

    finance = raw.get("financial_summary", {})
    prediction = raw.get("prediction", {})
    quality = raw.get("quality_metrics", {})

    rows = [
        ("\ucd1d\ub9e4\ucd9c", finance.get("total_revenue", 0), "KRW", "VAT \uc81c\uc678 \uae30\uc900"),
        ("\ucd1d\uc6d0\uac00", finance.get("total_cost", 0), "KRW", "BOM \uc6d0\uac00 \uae30\uc900"),
        ("\ub9e4\ucd9c\uc774\uc775", finance.get("gross_profit", 0), "KRW", "\ub9e4\ucd9c - \uc6d0\uac00"),
        ("\uc21c\uc774\uc775\ub960", finance.get("net_margin_rate", 0), "%", ""),
        ("\ucd1d \uc218\uc8fc\uac74\uc218", finance.get("total_orders", 0), "EA", ""),
        ("\ud3c9\uade0 \uc218\uc8fc\uae08\uc561", finance.get("avg_order_value", 0), "KRW", ""),
        ("\uc0dd\uc0b0 \ud488\uc9c8", "", "", ""),
        ("\uc591\ud488 \uc218\ub7c9", quality.get("total_good", 0), "EA", ""),
        ("\ubd88\ub7c9 \uc218\ub7c9", quality.get("total_bad", 0), "EA", ""),
        ("\uc218\uc728", quality.get("yield_rate", 0), "%", ""),
        ("\ubd88\ub7c9 \uae30\ud68c\ube44\uc6a9", raw.get("loss_cost", 0), "KRW", "defect_qty x unit_cost"),
        ("AI \uc608\uce21", "", "", ""),
        ("90\uc77c \uc608\uc0c1 \ub9e4\ucd9c", prediction.get("predicted_revenue", 0), "KRW", prediction.get("message", "")),
        ("90\uc77c \uc608\uc0c1 \uc774\uc775", prediction.get("predicted_profit", 0), "KRW", ""),
        ("\uc608\uc0c1 \uc774\uc775\ub960", prediction.get("margin_rate", 0), "%", ""),
    ]

    border = _thin_border()
    for row_index, row in enumerate(rows, 6):
        metric, value, unit, note = row
        for col_index, item in enumerate(row, 1):
            cell = ws.cell(row=row_index, column=col_index, value=item)
            cell.border = border
            if col_index == 1:
                cell.alignment = CL
                if unit == "" and note == "" and value == "":
                    cell.fill = SUBHDR_FILL
                    cell.font = HDR_FONT
                else:
                    cell.font = BODY_FONT
            elif col_index == 2 and isinstance(item, (int, float)):
                cell.alignment = CR
                cell.number_format = PCT if unit == "%" else (NUM_N if item < 0 else NUM)
                cell.font = NEG_FONT if item < 0 else BODY_FONT
            else:
                cell.alignment = CL
                cell.font = BODY_FONT
        _apply_alt(ws, row_index, 4)

    ws.conditional_formatting.add("B9", CellIsRule(operator="lessThan", formula=["0"], font=NEG_FONT, fill=NEG_FILL))
    ws.conditional_formatting.add(
        "B15",
        CellIsRule(operator="lessThan", formula=["95"], font=_font(10, bold=True, color=P["amber"]), fill=WARN_FILL),
    )

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 46


def _build_monthly(workbook: Workbook, raw: dict) -> None:
    ws = workbook.create_sheet(SHEET_MONTHLY)
    ws.sheet_properties.tabColor = P["navy_mid"]

    ws.merge_cells("A1:G1")
    ws["A1"] = "\uc6d4\ubcc4 \uc190\uc775"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 26

    cols = ["\uc6d4", "\ub9e4\ucd9c", "\uc6d0\uac00", "\uc774\uc775", "\uc774\uc775\ub960(%)", "\uc218\uc8fc\uac74\uc218", "\ud3c9\uade0\uc218\uc8fc\uae08\uc561"]
    for index, title in enumerate(cols, 1):
        ws.cell(row=2, column=index, value=title)
    _hdr_row(ws, 2, len(cols))

    monthly = _safe_rows(raw.get("monthly_performance"))
    last_data_row = 2 + len(monthly)

    border = _thin_border()
    for row_index, item in enumerate(monthly, 3):
        revenue = item.get("revenue", 0)
        cost = item.get("cost", 0)
        profit = item.get("profit", 0)
        orders = item.get("order_count", 0)
        margin = item.get("margin_rate", 0)
        avg_order = int(revenue / orders) if orders else 0
        values = [item.get("month", ""), revenue, cost, profit, margin, orders, avg_order]

        for col_index, value in enumerate(values, 1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.border = border
            if col_index == 1:
                cell.font = BODY_FONT
                cell.alignment = CC
            elif col_index == 5:
                cell.font = NEG_FONT if isinstance(value, (int, float)) and value < 0 else BODY_FONT
                cell.number_format = PCT
                cell.alignment = CR
            elif isinstance(value, (int, float)):
                cell.font = BODY_FONT
                cell.number_format = NUM_N if value < 0 else NUM
                cell.alignment = CR
            else:
                cell.font = BODY_FONT
                cell.alignment = CL
        _apply_alt(ws, row_index, len(cols))

    total_row = last_data_row + 1
    formulas = [
        "\ud569\uacc4",
        f"=SUM(B3:B{last_data_row})",
        f"=SUM(C3:C{last_data_row})",
        f"=SUM(D3:D{last_data_row})",
        f"=IFERROR(D{total_row}/B{total_row}*100,0)",
        f"=SUM(F3:F{last_data_row})",
        "",
    ]
    for col_index, value in enumerate(formulas, 1):
        cell = ws.cell(row=total_row, column=col_index, value=value)
        cell.fill = TOT_FILL
        cell.font = BOLD_FONT
        cell.border = border
        if col_index == 5:
            cell.number_format = PCT
            cell.alignment = CR
        elif col_index > 1:
            cell.number_format = NUM
            cell.alignment = CR

    if monthly:
        _add_table(ws, "tblMonthly", f"A2:G{last_data_row}", "TableStyleMedium9")

    cf_end = max(last_data_row, 5)
    _cf_margin(ws, "E", 3, cf_end)
    _cf_profit(ws, "D", 3, cf_end)
    _cf_databar(ws, "B", 3, cf_end, P["navy"])
    _auto_width(ws)


def _build_quarterly(workbook: Workbook, raw: dict) -> None:
    ws = workbook.create_sheet(SHEET_QUARTERLY)
    ws.sheet_properties.tabColor = "44546A"

    ws.merge_cells("A1:F1")
    ws["A1"] = "\ubd84\uae30\ubcc4 \uc2e4\uc801"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 26

    cols = ["\ubd84\uae30", "\ub9e4\ucd9c", "\uc6d0\uac00", "\uc774\uc775", "\uc774\uc775\ub960(%)", "\uc218\uc8fc\uac74\uc218"]
    for index, title in enumerate(cols, 1):
        ws.cell(row=2, column=index, value=title)
    _hdr_row(ws, 2, len(cols))

    quarterly = _safe_rows(raw.get("quarterly_chart"))
    last_data_row = 2 + len(quarterly)

    border = _thin_border()
    for row_index, item in enumerate(quarterly, 3):
        is_forecast = item.get("is_forecast", False)
        order_count = item.get("order_count", 0)
        values = [
            item.get("label", item.get("quarter", "")),
            item.get("revenue", 0),
            item.get("cost", 0),
            item.get("profit", 0),
            item.get("margin_rate", 0),
            "-" if is_forecast and order_count == 0 else order_count,
        ]

        for col_index, value in enumerate(values, 1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.border = border
            cell.font = FC_FONT if is_forecast else BODY_FONT
            if col_index == 1:
                cell.alignment = CC
            elif col_index == 5:
                cell.number_format = PCT
                cell.alignment = CR
            elif isinstance(value, (int, float)):
                cell.number_format = NUM_N if value < 0 else NUM
                cell.alignment = CR
            else:
                cell.alignment = CL
        _apply_alt(ws, row_index, len(cols))

    total_row = last_data_row + 1
    formulas = [
        "\ud569\uacc4",
        f"=SUM(B3:B{last_data_row})",
        f"=SUM(C3:C{last_data_row})",
        f"=SUM(D3:D{last_data_row})",
        f"=IFERROR(D{total_row}/B{total_row}*100,0)",
        f"=SUM(F3:F{last_data_row})",
    ]
    for col_index, value in enumerate(formulas, 1):
        cell = ws.cell(row=total_row, column=col_index, value=value)
        cell.fill = TOT_FILL
        cell.font = BOLD_FONT
        cell.border = border
        if col_index == 5:
            cell.number_format = PCT
            cell.alignment = CR
        elif col_index > 1:
            cell.number_format = NUM
            cell.alignment = CR

    if quarterly:
        _add_table(ws, "tblQuarterly", f"A2:F{last_data_row}", "TableStyleMedium2")

    cf_end = max(last_data_row, 5)
    _cf_margin(ws, "E", 3, cf_end)
    _cf_profit(ws, "D", 3, cf_end)
    _cf_databar(ws, "B", 3, cf_end, P["navy"])
    _auto_width(ws)


def _build_forecast(workbook: Workbook, raw: dict) -> None:
    ws = workbook.create_sheet(SHEET_FORECAST)
    ws.sheet_properties.tabColor = P["amber"]

    ws.merge_cells("A1:E1")
    ws["A1"] = "AI \uc7a5\uae30 \uc608\uce21"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 26

    cols = ["\uc6d4", "\uc608\uc0c1 \ub9e4\ucd9c", "\uc608\uc0c1 \uc6d0\uac00", "\uc608\uc0c1 \uc774\uc775", "\uc774\uc775\ub960(%)"]
    for index, title in enumerate(cols, 1):
        ws.cell(row=2, column=index, value=title)
    _hdr_row(ws, 2, len(cols))

    forecast = _safe_rows(raw.get("long_term_forecast"))
    last_data_row = 2 + len(forecast)

    border = _thin_border()
    for row_index, item in enumerate(forecast, 3):
        revenue = item.get("predicted_revenue", 0)
        cost = item.get("predicted_cost", 0)
        profit = item.get("predicted_profit", 0)
        margin = item.get("margin_rate", round((profit / revenue) * 100, 2) if revenue else 0)
        values = [item.get("month", ""), revenue, cost, profit, margin]

        for col_index, value in enumerate(values, 1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.border = border
            cell.font = FC_FONT
            if col_index == 1:
                cell.alignment = CC
            elif col_index == 5:
                cell.number_format = PCT
                cell.alignment = CR
            elif isinstance(value, (int, float)):
                cell.number_format = NUM
                cell.alignment = CR
        _apply_alt(ws, row_index, len(cols))

    total_row = last_data_row + 1
    formulas = [
        "\ud569\uacc4",
        f"=SUM(B3:B{last_data_row})",
        f"=SUM(C3:C{last_data_row})",
        f"=SUM(D3:D{last_data_row})",
        f"=IFERROR(D{total_row}/B{total_row}*100,0)",
    ]
    for col_index, value in enumerate(formulas, 1):
        cell = ws.cell(row=total_row, column=col_index, value=value)
        cell.fill = TOT_FILL
        cell.font = BOLD_FONT
        cell.border = border
        if col_index == 5:
            cell.number_format = PCT
            cell.alignment = CR
        elif col_index > 1:
            cell.number_format = NUM
            cell.alignment = CR

    if forecast:
        _add_table(ws, "tblForecast", f"A2:E{last_data_row}", "TableStyleMedium7")

    cf_end = max(last_data_row, 5)
    ws.conditional_formatting.add(
        f"E3:E{cf_end}",
        ColorScaleRule(
            start_type="min",
            start_color="FFFFDADA",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFFFF7CE",
            end_type="max",
            end_color="FFD7F0E6",
        ),
    )
    _cf_databar(ws, "B", 3, cf_end, P["amber"])
    _auto_width(ws)


def _build_quality(workbook: Workbook, raw: dict) -> None:
    ws = workbook.create_sheet(SHEET_QUALITY)
    ws.sheet_properties.tabColor = P["green"]

    ws.merge_cells("A1:C1")
    ws["A1"] = "\ud488\uc9c8 \ubd84\uc11d"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 26

    quality = raw.get("quality_metrics", {})
    rows = [
        ("\uad6c\ubd84", "\uc218\ub7c9", "\ube44\uc728(%)"),
        ("\uc591\ud488", quality.get("total_good", 0), "=B3/(B3+B4)*100"),
        ("\ubd88\ub7c9", quality.get("total_bad", 0), "=B4/(B3+B4)*100"),
        ("\ud569\uacc4", "=B3+B4", "100.00"),
        ("\uc218\uc728", "", quality.get("yield_rate", 0)),
        ("\ubd88\ub7c9 \uae30\ud68c\ube44\uc6a9", raw.get("loss_cost", 0), ""),
    ]

    border = _thin_border()
    for row_index, row in enumerate(rows, 2):
        for col_index, value in enumerate(row, 1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.border = border
            if row_index == 2:
                cell.fill = HDR_FILL
                cell.font = HDR_FONT
                cell.alignment = CC
            else:
                cell.font = BODY_FONT
                if col_index == 3 and row_index > 2:
                    cell.number_format = PCT
                    cell.alignment = CR
                elif isinstance(value, (int, float)):
                    cell.number_format = NUM
                    cell.alignment = CR
                else:
                    cell.alignment = CL

    ws.conditional_formatting.add(
        "A4:C4",
        FormulaRule(formula=['$A4="\ubd88\ub7c9"'], fill=NEG_FILL, font=_font(10, bold=True, color=P["red"])),
    )
    ws.conditional_formatting.add(
        "C6",
        CellIsRule(operator="lessThan", formula=["95"], font=_font(10, bold=True, color=P["amber"]), fill=WARN_FILL),
    )

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14


def _build_cancel(workbook: Workbook, raw: dict) -> None:
    ws = workbook.create_sheet(SHEET_CANCEL)
    ws.sheet_properties.tabColor = P["red"]

    ws.merge_cells("A1:C1")
    ws["A1"] = "\uc218\uc8fc \ucde8\uc18c \ubd84\uc11d"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 26

    cancel = raw.get("cancel_analysis", {})
    rows = [
        ("\uad6c\ubd84", "\uac74\uc218", "\ube44\uc728(%)"),
        ("\uc804\uccb4 \uc218\uc8fc", cancel.get("total_orders", 0), "100.00"),
        ("\uc815\uc0c1 \uc218\uc8fc", cancel.get("done_orders", 0), "=B4/B3*100"),
        ("\ucde8\uc18c \uc218\uc8fc", cancel.get("cancel_orders", 0), "=B5/B3*100"),
        ("\ucde8\uc18c \uc190\uc2e4\uc561", cancel.get("cancel_amount", 0), ""),
        ("\ucde8\uc18c\uc728", "", cancel.get("cancel_rate", 0)),
    ]

    border = _thin_border()
    for row_index, row in enumerate(rows, 2):
        for col_index, value in enumerate(row, 1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.border = border
            if row_index == 2:
                cell.fill = HDR_FILL
                cell.font = HDR_FONT
                cell.alignment = CC
            else:
                cell.font = BODY_FONT
                if col_index == 3:
                    cell.number_format = PCT
                    cell.alignment = CR
                elif isinstance(value, (int, float)) and value != 0:
                    cell.number_format = NUM
                    cell.alignment = CR
                else:
                    cell.alignment = CL
        if row_index % 2 == 0:
            for col_index in range(1, 4):
                cell = ws.cell(row=row_index, column=col_index)
                if cell.fill.fgColor.type != "rgb" or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                    cell.fill = ALT_FILL

    ws.conditional_formatting.add(
        "B5",
        CellIsRule(operator="greaterThan", formula=["10"], font=_font(10, bold=True, color=P["red"]), fill=NEG_FILL),
    )
    ws.conditional_formatting.add(
        "C7",
        CellIsRule(operator="greaterThan", formula=["5"], font=_font(10, bold=True, color=P["amber"]), fill=WARN_FILL),
    )

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
